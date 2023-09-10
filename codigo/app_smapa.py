import os
import time
import shutil
import re

import pyautogui
from openpyxl import load_workbook
import fitz
import glob

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

class Scraper_Smapa():

    def __init__(self,url, email, password, driver_path):
        print(url, email, password, driver_path)
        self.url = url
        self.email = email
        self.password = password
        self.driver_path = driver_path

    def wait(self, seconds):
        return WebDriverWait(self.driver, seconds)

    def close(self):
        self.driver.close()
        self.driver = None

    def quit(self):
        self.driver.quit()
        self.driver = None    

    def diccionario(self,mes_texto):
        dic = {'ene.':'01','feb.':'02','mar.':'03','abr.':'04','may.':'05','jun.':'06','jul.':'07',
               'ago.':'08','sep.':'09','oct.':'10','nov.':'11','dic.':'12'}

        mes_oficial = dic[mes_texto]
        
        return mes_oficial

    def login(self):
        
        driver_exe = '.domain\\chromedriver.exe'
        credencials = '.\\config\\credenciales.xlsx'

        print('Entrando en la funcion login...')
        print('----------------------------------------------------------------------')
        
        #Seteo variables
        email = self.email
        url = self.url
        driver_path = self.driver_path
        password  = self.password
        
        options = webdriver.ChromeOptions()
        options.add_experimental_option('prefs', {
        "download.default_directory": "C:\\roda\\smapa-portal\\input\\", #Change default directory for downloads
        "download.prompt_for_download": False, #To auto download the file
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True, #It will not show PDF directly in chrome
        })  
        
        self.driver = webdriver.Chrome(driver_path, options=options)
        self.driver.get(url)
        self.driver.maximize_window()
        self.driver.implicitly_wait(30)

        # Controlar evento alert() Notificacion
        try:
            alert = WebDriverWait(self.driver, 35).until(EC.alert_is_present())
            alert = Alert(self.driver)
            alert.accept()
                    
        except:
            print("No se encontró ninguna alerta.")  

        #Botones ID que utilizaremos para logear
        selector_ingreso_cuenta = 'Mail'
        selector_password_input = 'Password'
        selector_login_button = 'btn-primary'

        # Seleccionar campo cuenta
        intentos = 0
        reintentar = True
        while (reintentar):
            try:
                print('Try en la funcion login para el campo cuenta...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_cuenta= self.driver.find_element(By.NAME, selector_ingreso_cuenta)
                element_cuenta.clear()
                element_cuenta.click()
                element_cuenta.send_keys(email)
                reintentar = False
            except:    
                print('Exception en la funcion click campo cuenta')
                print('----------------------------------------------------------------------')
                reintentar = intentos <= 3                
                
        # Seleccionar campo clave y setear clave
        intentos = 0
        reintentar = True
        while (reintentar):
            try:
                print('Try en la funcion click para el campo rut...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_password = self.driver.find_element(By.NAME, selector_password_input)
                element_password.clear()
                element_password.click()
                element_password.send_keys(password)
                reintentar = False
            except:    
                print('Exception en la funcion click rut cliente')
                print('----------------------------------------------------------------------')
                reintentar = intentos <= 3

        # Seleccionar boton y hacer en boton ingresar
        intentos = 0
        reintentar = True
        while (reintentar):
            try:
                print('Try en la funcion click_element_xpath para el boton ingresar...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_button_ingresar = self.driver.find_element(By.CLASS_NAME, selector_login_button)
                element_button_ingresar.click()
                reintentar = False
            except:    
                print('Exception en la funcion click_element_xpath')
                print('----------------------------------------------------------------------')
                reintentar = intentos <= 3
    
    def scrapping_smapa(self,sociedad,limite):

        while sociedad < limite:
        
            #Buscamos la tabla que contiene las filas a buscar
            intento_sociedades = 0
            reintentar_sociedades = True
            while (reintentar_sociedades):
                try:
                    time.sleep(5)
                    elemento = self.driver.find_element(By.XPATH,f"/html/body/div[1]/div[2]/main/div/div/div/div/div/div[2]/div/div/div[2]/div/div[{sociedad}]/div[1]/a")
                    print(f'la sociedad es la numero {sociedad}')
                    numero_sociedad = elemento.text
                    elemento.click()
                    reintentar_sociedades = False
                except:    
                    print('menu sociedades no esta disponible')
                    print('----------------------------------------------------------------------')
                    self.driver.execute_script("window.scrollBy(0, 700);")
                    reintentar_sociedades = intento_sociedades <= 5 

            #Aqui, una vez que ya entramos a la sociedad, obtenemos la cantidad de boletas para comenzar descarga. Sera un bucle separado
            filas_tabla = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.TAG_NAME,"tr"))) 
            
            #Buscamos el largo de las descargas para utilizarlo como limite
            filas = self.driver.find_elements(By.TAG_NAME,"tr")
            
            cantidad_facturas =len(filas)
            print(f'cantidad de facturas es{cantidad_facturas}')
            print('----------------------------------------------------------')
            
            i=1
            while i < cantidad_facturas:

                #extraer el numero de factura
                intento = 0
                factura = True
                while (factura):
                    try:
                        time.sleep(5)
                        boton_factura = self.driver.find_element(By.XPATH,f'/html/body/div[1]/div[2]/main/div/div/div/div/div/div[3]/div/div[2]/div/div[2]/div/div/div/table/tbody/tr[{i}]/td[1]')
                        numero_fact = boton_factura.text                            
                        lista = numero_fact.split()
                        factura_oficial = lista[2]
                        print(factura_oficial)
                        factura = False
                    except:    
                        print('Boton de descarga no se encuentra disponible....')
                        print('----------------------------------------------------------------------')
                        factura = intento <= 5    
                
                #Buscamos el boton de descarga en cada una de las filas
                intento = 0
                descarga = True
                while (descarga):
                    try:
                        time.sleep(5)
                        boton_descarga = self.driver.find_element(By.XPATH,f'/html/body/div[1]/div[2]/main/div/div/div/div/div/div[3]/div/div[2]/div/div[2]/div/div/div/table/tbody/tr[{i}]/td[6]/div/button')  
                        boton_descarga.click()
                        descarga = False
                    except:    
                        print('Boton de descarga no se encuentra disponible....')
                        print('----------------------------------------------------------------------')
                        descarga = intento <= 5      

                # Esperar a que se abra la ventana emergente
                intentos = 0
                reintentar_ventana = True
                while (reintentar_ventana):
                    try:
                        # Obtiene el identificador de la ventana actual
                        current_window = self.driver.current_window_handle
                        print('Ventana principal: ', current_window)
                        print('----------------------------------------------------------------------')
                        print('Try en la funcion manejo de ventanas abiertas...', intentos)
                        print('----------------------------------------------------------------------')
                        intentos += 1
                        window_handles_all = self.driver.window_handles
                        reintentar_ventana = False
                            
                    except:    
                        print('Exception en la funcion manejo de ventanas abiertas')
                        print('----------------------------------------------------------------------')
                        print('----------------------------------------------------------------------')
                        time.sleep(60) #espera para que cargue ventana emergente
                                    
                        reintentar_ventana = intentos <= 3
                                    
                # Obtiene los identificadores de las ventanas abiertas
                self.driver.implicitly_wait(35)
                print('Ventanas abiertas: ', window_handles_all, len(window_handles_all))
                print('----------------------------------------------------------------------')            
                                    
                # Cambiar al manejo de la ventana emergente
                for window_handle in window_handles_all:
                    if window_handle != current_window:
                        self.driver.switch_to.window(window_handle)
                        print('Ventana emergente: ', window_handle)
                        print('----------------------------------------------------------------------')
                        break
                
                
                #Revisar si este es el primer documento o no
                #print(os.listdir(folder_path))
                if i ==1 and sociedad==1:
                    hay_archivos = False
                elif i !=1 and sociedad ==1:
                    hay_archivos = True
                elif i ==1 and sociedad !=1:
                    hay_archivos = True
                elif i !=1 and sociedad !=1:
                    hay_archivos = True
                
                # Esperar hasta que el elemento esté presente en la página para descargar
                time.sleep(25) 
                pyautogui.click(500, 500)
                
                #AMBAS OPCIONES FUNCIONAN
                actions = ActionChains(self.driver)
                actions.send_keys(Keys.RETURN).perform()

                # Cerrar la ventana emergente
                time.sleep(5)
                self.driver.close()
                
                print(f'la pasada numero {i} entrega hay archivos igual a {hay_archivos}')
                
                #Ruta de descarga
                folder_path = './input/'
                
                if hay_archivos == False:
                
                    #Buscamos el archivo que se haya descargado que comience con el nombre boleta para poder moverlo a la carpeta input
                    for filename in os.listdir(folder_path):
                        if filename.endswith('.pdf'):
                            nombre_archivo = filename
                            os.rename(folder_path+nombre_archivo,folder_path+f'soc_{numero_sociedad}_fac_{factura_oficial}.pdf')
                
                elif hay_archivos == True:
                    
                    for filename in os.listdir(folder_path):
                        if not filename.startswith('soc_') and filename.endswith('.pdf'):
                            nombre_archivo = filename
                            os.rename(folder_path+nombre_archivo,folder_path+f'soc_{numero_sociedad}_fac_{factura_oficial}.pdf')   

                # Cambiar de nuevo al manejo de ventana principal
                self.driver.switch_to.window(current_window)
                print('Cual ventana es: ', current_window)
                print('----------------------------------------------------------------------')
                        
                time.sleep(10)
                
                print('pasamos al siguiente archivo')
                i+=1
            
            sociedad+=1  
            self.driver.execute_script("window.history.go(-1)")
            print('Pasamos a la siguiente sociedad')

    def archivos(self):
        
        folder_path = './input/'
        output_path = './output/'
        
        #Revisamos si hay archivos pdf en la carpeta input
        archivos_pdf = glob.glob(os.path.join(folder_path, '*.pdf'))

        #Si no encuentra archivos es porque no se realizo la ejecucion correcta y hay que mandar mail
        if not archivos_pdf:
            print(f'No se encontraron archivos PDF en la carpeta "{folder_path}".')
        else:
            #Si encuentra me entregara todos los documentos con los que trabajaremos
            print(f'Se encontraron los siguientes archivos PDF en la carpeta "{folder_path}":')
            
            for archivo in archivos_pdf:
                nombre_oficial = archivo.replace('./input','')
                
                print(archivo)
                            
                with fitz.open(archivo) as pdf_documento:
                    texto_completo = ''

                    for pagina_num in range(pdf_documento.page_count):
                        pagina = pdf_documento.load_page(pagina_num)
                        texto_completo += pagina.get_text()
                    
                    lista_limpia = [elemento.strip() for elemento in texto_completo.split('\n')]

                    #TABLA 1
                    
                    #Posicion 1
                    factura_elec = lista_limpia[15]
                    
                    #Posicion 2
                    direccion_instalacion = lista_limpia[20]
                    
                    #Posicion 3
                    giro_bruto = lista_limpia[1]
                    partes = giro_bruto.split("GIRO: ")
                    giro = partes[1]
                    
                    #Posicion 4
                    id_servicio_bruto = lista_limpia[23]
                    partes = id_servicio_bruto.split("ID Servicio: ")
                    id_servicio = partes[1]
                    
                    #TABLA 2
                    
                    #Posicion 5
                    fecha_emision_bruto = lista_limpia[25]
                    partes = fecha_emision_bruto.split(":")
                    fecha_emision_bru = partes[1].strip()
                    partes = fecha_emision_bru.split()
                    dia = partes[0]
                    mes = self.diccionario(str(partes[1]))
                    año = partes[2]
                    
                    fecha_emision = dia+'-'+mes+'-'+año
                    
                    #Posicion 6
                    fecha_vencimiento_bruto = lista_limpia[26]
                    partes = fecha_vencimiento_bruto.split(":")
                    fecha_vencimiento_bruto = partes[1].strip()
                    if fecha_vencimiento_bruto == 'CORTE EN TRAMITE':
                        fecha_vencimiento = fecha_vencimiento_bruto
                    elif fecha_vencimiento_bruto != 'CORTE EN TRAMITE':
                        partes = fecha_vencimiento_bruto.split()
                        dia = partes[0]
                        mes = self.diccionario(str(partes[1]))
                        año = partes[2]
                        
                        fecha_vencimiento = dia+'-'+mes+'-'+año  
                    
                    #Posicion 7
                    direccion_central_bruto = lista_limpia[27]
                    partes = direccion_central_bruto.split(": ")
                    direccion_central_casi = partes[1].strip()
                    direccion_central = direccion_central_casi.replace(',','')
                    
                    #TABLA 3
                    
                    #Posicion 8
                    elemento_a_buscar = 'CARGO FIJO'
                    try:
                        posicion_8 = lista_limpia.index(elemento_a_buscar)
                        total_cargo_fijo_b = lista_limpia[posicion_8+1]
                        total_cargo_fijo = total_cargo_fijo_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        total_cargo_fijo = 0
 
                    #Posicion 9
                    elemento_a_buscar = 'CONSUMO AGUA POTABLE PUNTA'
                    try:
                        posicion_9 = lista_limpia.index(elemento_a_buscar)
                        cantidad_agua_potable = lista_limpia[posicion_9+1]
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_agua_potable = 0
                    
                    #Posicion 10
                    elemento_a_buscar = 'CONSUMO AGUA POTABLE PUNTA'
                    try:
                        posicion_10 = lista_limpia.index(elemento_a_buscar)
                        precio_agua_potable_b = lista_limpia[posicion_10+2]
                        precio_agua_potable = precio_agua_potable_b.replace('$ ','')
                    except:
                        print('elemento no se encuentra disponible')
                        precio_agua_potable = 0
                    
                    #Posicion 11
                    elemento_a_buscar = 'CONSUMO AGUA POTABLE PUNTA'
                    try:
                        posicion_11 = lista_limpia.index(elemento_a_buscar)
                        total_agua_potable_b = lista_limpia[posicion_11+3]
                        total_agua_potable  = total_agua_potable_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        total_agua_potable = 0
                    
                    #Posicion 12
                    elemento_a_buscar = 'SOBRECONSUMO AGUA POTABLE'
                    try:
                        posicion_12 = lista_limpia.index(elemento_a_buscar)
                        cantidad_sobreconsumo = lista_limpia[posicion_12+1]
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_sobreconsumo = 0
                    
                    #Posicion 13
                    elemento_a_buscar = 'SOBRECONSUMO AGUA POTABLE'
                    try:
                        posicion_13 = lista_limpia.index(elemento_a_buscar)
                        precio_sobreconsumo_b = lista_limpia[posicion_13+2]
                        precio_sobreconsumo = precio_sobreconsumo_b.replace('$ ','')
                    except:
                        print('elemento no se encuentra disponible')
                        precio_sobreconsumo = 0
                    
                    #Posicion 14
                    elemento_a_buscar = 'SOBRECONSUMO AGUA POTABLE'
                    try:
                        posicion_14 = lista_limpia.index(elemento_a_buscar)
                        total_sobreconsumo_b = lista_limpia[posicion_14+3]
                        total_sobreconsumo = total_sobreconsumo_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        total_sobreconsumo = 0
                    
                    #Posicion 15
                    elemento_a_buscar = 'ALCANTARILLADO S/TRATA'
                    try:
                        posicion_15 = lista_limpia.index(elemento_a_buscar)
                        cantidad_alcantarillado = lista_limpia[posicion_15+1]
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_alcantarillado = 0
                    
                    #Posicion 16
                    elemento_a_buscar = 'ALCANTARILLADO S/TRATA'
                    try:
                        posicion_16 = lista_limpia.index(elemento_a_buscar)
                        precio_alcantarillado_b = lista_limpia[posicion_16+2]
                        precio_alcantarillado = precio_alcantarillado_b.replace('$ ','')
                    except:
                        print('elemento no se encuentra disponible')
                        precio_alcantarillado = 0
                    
                    #Posicion 17
                    elemento_a_buscar = 'ALCANTARILLADO S/TRATA'
                    try:
                        posicion_17 = lista_limpia.index(elemento_a_buscar)
                        total_alcantarillado = lista_limpia[posicion_17+3]
                    except:
                        print('elemento no se encuentra disponible')
                        total_alcantarillado = 0
                    
                    #Posicion 18
                    elemento_a_buscar = '* TRATAM. AGUAS SERVIDAS'
                    try:
                        posicion_18 = lista_limpia.index(elemento_a_buscar)
                        cantidad_tratam_aguas_servidas = lista_limpia[posicion_18+1]
                    except:
                        print('elemento no se encuentra disponible')
                        cantidad_tratam_aguas_servidas = 0
                    
                    #Posicion 19
                    elemento_a_buscar = '* TRATAM. AGUAS SERVIDAS'
                    try:
                        posicion_19 = lista_limpia.index(elemento_a_buscar)
                        precio_tratam_aguas_servidas_b = lista_limpia[posicion_19+2]
                        precio_tratam_aguas_servidas = precio_tratam_aguas_servidas_b.replace('$ ','')
                    except:
                        print('elemento no se encuentra disponible')
                        precio_tratam_aguas_servidas = 0
                    
                    #Posicion 20
                    elemento_a_buscar = '* TRATAM. AGUAS SERVIDAS'
                    try:
                        posicion_20 = lista_limpia.index(elemento_a_buscar)
                        total_tratam_aguas_servidas_b = lista_limpia[posicion_20+3]
                        total_tratam_aguas_servidas = total_tratam_aguas_servidas_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        total_tratam_aguas_servidas = 0
                
                    #TABLA 4
                
                    #Posicion 21
                    elemento_a_buscar = 'OBSERVACIONES'
                    try:
                        posicion_21 = lista_limpia.index(elemento_a_buscar)
                        fecha_tarifa_publicada_bruto = lista_limpia[posicion_21+1]
                        indice_ultimo_espacio = fecha_tarifa_publicada_bruto.rfind(" ")
                        fecha_tarifa_publicada_b = fecha_tarifa_publicada_bruto[indice_ultimo_espacio + 1:]
                        fecha_tarifa_publicada = fecha_tarifa_publicada_b.replace('/','-')
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_tarifa_publicada = '-'

                    #Posicion 22
                    elemento_a_buscar = 'OBSERVACIONES'
                    try:
                        posicion_22 = lista_limpia.index(elemento_a_buscar)
                        cargo_variable_AP_bruto = lista_limpia[posicion_22+2]
                        indice_signo_dolar = cargo_variable_AP_bruto.index("$")
                        cargo_variable_AP_b = cargo_variable_AP_bruto[indice_signo_dolar:]
                        cargo_variable_AP = cargo_variable_AP_b.replace('$ ','')
                    except:
                        print('elemento no se encuentra disponible')
                        cargo_variable_AP = 0
                    
                    #Posicion 23
                    elemento_a_buscar = 'OBSERVACIONES'
                    try:
                        posicion_23 = lista_limpia.index(elemento_a_buscar)
                        cargo_variable_AP_punta_bruto = lista_limpia[posicion_23+3]
                        indice_signo_dolar = cargo_variable_AP_punta_bruto.index("$")
                        cargo_variable_AP_punta_b = cargo_variable_AP_punta_bruto[indice_signo_dolar:]
                        cargo_variable_AP_punta = cargo_variable_AP_punta_b.replace('$ ','')
                    except:
                        print('elemento no se encuentra disponible')
                        cargo_variable_AP_punta = 0
                    
                    #Posicion 24
                    elemento_a_buscar = 'OBSERVACIONES'
                    try:
                        posicion_24 = lista_limpia.index(elemento_a_buscar)
                        cargo_variable_AL_bruto = lista_limpia[posicion_24+4]
                        indice_signo_dolar = cargo_variable_AL_bruto.index("$")
                        cargo_variable_AL_b = cargo_variable_AL_bruto[indice_signo_dolar:]
                        cargo_variable_AL = cargo_variable_AL_b.replace('$ ','')
                    except:
                        print('elemento no se encuentra disponible')
                        cargo_variable_AL = 0
                    
                    #Posicion 25
                    elemento_a_buscar = 'OBSERVACIONES'
                    try:
                        posicion_25 = lista_limpia.index(elemento_a_buscar)
                        cargo_variable_sobreconsumo_bruto = lista_limpia[posicion_25+5]
                        indice_signo_dolar = cargo_variable_sobreconsumo_bruto.index("$")
                        cargo_variable_sobreconsumo_b = cargo_variable_sobreconsumo_bruto[indice_signo_dolar:]
                        cargo_variable_sobreconsumo = cargo_variable_sobreconsumo_b.replace('$ ','')
                    except:
                        print('elemento no se encuentra disponible')
                        cargo_variable_sobreconsumo = 0
                        
                    #Posicion 26
                    elemento_a_buscar = '* Corte llave de paso:'
                    try:
                        posicion_26 = lista_limpia.index(elemento_a_buscar)
                        corte_llave_paso_b = lista_limpia[posicion_26+1]
                        corte_llave_paso = corte_llave_paso_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        corte_llave_paso = 0
                    
                    #Posicion 27
                    elemento_a_buscar = '* Reposición llave de paso:'
                    try:
                        posicion_27 = lista_limpia.index(elemento_a_buscar)
                        reposicion_llave_b = lista_limpia[posicion_27+1]
                        reposicion_llave = reposicion_llave_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        reposicion_llave = 0
                    
                    #TABLA 5
                 
                    #Posicion 28
                    elemento_a_buscar = 'Monto Neto'
                    try:
                        posicion_28 = lista_limpia.index(elemento_a_buscar)
                        monto_neto_b = lista_limpia[posicion_28+1]
                        monto_neto = monto_neto_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        monto_neto = 0
                    
                    #Posicion 29
                    elemento_a_buscar = 'IVA'
                    try:
                        posicion_29 = lista_limpia.index(elemento_a_buscar)
                        iva_b = lista_limpia[posicion_29+1]
                        iva = iva_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        iva = 0
                    
                    #Posicion 30
                    elemento_a_buscar = 'Total Mes'
                    try:
                        posicion_30 = lista_limpia.index(elemento_a_buscar)
                        total_mes_b = lista_limpia[posicion_30+1]
                        total_mes = total_mes_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        total_mes = 0
                    
                    #Posicion 31
                    elemento_a_buscar = 'Saldo Anterior'
                    try:
                        posicion_31 = lista_limpia.index(elemento_a_buscar)
                        saldo_anterior_b = lista_limpia[posicion_31+1]
                        saldo_anterior = saldo_anterior_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        saldo_anterior = 0

                    #Posicion 32
                    elemento_a_buscar = 'Total A Pagar'
                    try:
                        posicion_32 = lista_limpia.index(elemento_a_buscar)
                        total_a_pagar_b = lista_limpia[posicion_32+1]
                        total_a_pagar = total_a_pagar_b.replace('$ ','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        total_a_pagar = 0

                    #SIGUIENTE PAGINA
                    #TABLA 6
                    
                    #Posicion 33
                    elemento_a_buscar = 'RUTA:'
                    try:
                        posicion_33 = lista_limpia.index(elemento_a_buscar)
                        ruta = lista_limpia[posicion_33+1]
                    except:
                        print('elemento no se encuentra disponible')
                        ruta = '-'
                        
                    #Posicion 34
                    elemento_a_buscar = 'TIPO SERVICIO:'
                    try:
                        posicion_34 = lista_limpia.index(elemento_a_buscar)
                        tipo_servicio = lista_limpia[posicion_34+1]
                    except:
                        print('elemento no se encuentra disponible')
                        tipo_servicio = '-'
                        
                    #Posicion 35
                    elemento_a_buscar = 'SUMINISTRO:'
                    try:
                        posicion_35 = lista_limpia.index(elemento_a_buscar)
                        suministro_1 = lista_limpia[posicion_35+1]
                        suministro_2 = lista_limpia[posicion_35+2]
                        suministro = suministro_1 + ' ' + suministro_2
                    except:
                        print('elemento no se encuentra disponible')
                        suministro = '-'
                        
                    #Posicion 36
                    elemento_a_buscar = 'GRUPO TARIF:'
                    try:
                        posicion_36 = lista_limpia.index(elemento_a_buscar)
                        grupo_tarifario = lista_limpia[posicion_36+1]
                    except:
                        print('elemento no se encuentra disponible')
                        grupo_tarifario = '-'
                        
                    #Posicion 37
                    elemento_a_buscar = 'MEDIDOR GRAL'
                    try:
                        posicion_37 = lista_limpia.index(elemento_a_buscar)
                        medidor_general = lista_limpia[posicion_37+1]
                    except:
                        print('elemento no se encuentra disponible')
                        medidor_general = '-'
                        
                    #Posicion 38
                    elemento_a_buscar = 'NUMERO MEDIDOR:'
                    try:
                        posicion_38 = lista_limpia.index(elemento_a_buscar)
                        n_medidor = lista_limpia[posicion_38+1]
                    except:
                        print('elemento no se encuentra disponible')
                        n_medidor = '-'
                        
                    #Posicion 39
                    elemento_a_buscar = 'DIAMETRO:'
                    try:
                        posicion_39 = lista_limpia.index(elemento_a_buscar)
                        diametro = lista_limpia[posicion_39+1]
                    except:
                        print('elemento no se encuentra disponible')
                        diametro = '-'
                        
                    #Posicion 40
                    elemento_a_buscar = 'PROXIMA LECTURA:'
                    try:
                        posicion_40 = lista_limpia.index(elemento_a_buscar)
                        proxima_lectura_b = lista_limpia[posicion_40+1]
                        partes = proxima_lectura_b.split()
                        dia = partes[0]
                        mes = self.diccionario(str(partes[1]))
                        año = partes[2]
                        proxima_lectura = dia+'-'+mes+'-'+año  
                        
                    except:
                        print('elemento no se encuentra disponible')
                        proxima_lectura = '-'
                    
                    #TABLA 7
                    
                    #Posicion 41
                    elemento_a_buscar = 'METODO CONSUMO:'
                    try:
                        posicion_41 = lista_limpia.index(elemento_a_buscar)
                        metodo_consumo = lista_limpia[posicion_41+1]
                    except:
                        print('elemento no se encuentra disponible')
                        metodo_consumo = '-'
                    
                    #Posicion 42 y 43
                    elemento_a_buscar = 'LECTURA ACTUAL'
                    for idx, elemento in enumerate(lista_limpia):
                        if elemento_a_buscar in elemento:
                            lector_actual_valor_bruto= lista_limpia[idx] #Este valor es lectura actual
                            
                            if 'm3' in lista_limpia[idx+1]:
                                de_m3_match = re.search(r'(.+m3)', lista_limpia[idx+1])
                                lector_actual_valor_b = de_m3_match.group(1)
                                lector_actual_valor = lector_actual_valor_b.replace(".", "").replace(" m3", "")
                                try:
                                    despues_de_m3 = re.search(r'm3\s+(.*)', lista_limpia[idx+1])
                                    lector_actual_fecha_b = despues_de_m3.group(1)
                                    partes = lector_actual_fecha_b.split()
                                    dia = partes[0]
                                    mes = self.diccionario(str(partes[1]))
                                    año = partes[2]
                                    lector_actual_fecha = dia+'-'+mes+'-'+año 
                                except:
                                    print('no hay fecha disponible')
                                    lector_actual_fecha = ''
                                    lector_actual_valor = 0
                    
                    #Posicion 44 y 45
                    texto_a_verificar = 'LECTURA ANTERIOR'
                    posicion_44 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_44 = idx
                            lectura_anterior_valor_bruto = lista_limpia[posicion_44]#Este valor es lectura anterior
                            despues_de_anterior = re.search(r'ANTERIOR\s+(.*m3)', lectura_anterior_valor_bruto)
                            if despues_de_anterior == None:
                                despues_de_anterior_match = re.search(r'ANTERIOR\s+(.*)', lectura_anterior_valor_bruto)
                                cantidad_sola = despues_de_anterior_match.group(1)
                                m3_solo = lista_limpia[posicion_44+1]
                                lectura_anterior_valor = cantidad_sola+' '+m3_solo
                                lectura_anterior_valor_real = lectura_anterior_valor.replace('.','').replace(" m3",'')
                                lector_anterior_fecha= lista_limpia[posicion_44+2]
                            
                            elif despues_de_anterior != None:
                                lectura_anterior_valor = despues_de_anterior.group(1)
                                lectura_anterior_valor_real = lectura_anterior_valor.replace('.','').replace(" m3",'')
                                try:
                                    despues_de_m3_match = re.search(r'm3\s+(.*)', lectura_anterior_valor_bruto)
                                    lector_anterior_fecha_b = despues_de_m3_match.group(1)
                                    partes = lector_anterior_fecha_b.split()
                                    dia = partes[0]
                                    mes = self.diccionario(str(partes[1]))
                                    año = partes[2]
                                    lector_anterior_fecha = dia+'-'+mes+'-'+año 
                                    break
                                except:
                                    lector_anterior_fecha_b= lista_limpia[posicion_44+1]
                                    partes = lector_anterior_fecha_b.split()
                                    dia = partes[0]
                                    mes = self.diccionario(str(partes[1]))
                                    año = partes[2]
                                    lector_anterior_fecha = dia+'-'+mes+'-'+año 
                                    break

                    #Posicion 46
                    elemento_a_buscar = 'CLAVE ACTUAL:'
                    try:
                        posicion_46 = lista_limpia.index(elemento_a_buscar)
                        clave_actual = lista_limpia[posicion_46+1]
                    except:
                        print('elemento no se encuentra disponible')
                        clave_actual = '-'
                    
                    #Posicion 47
                    elemento_a_buscar = 'CONSUMO'
                    try:
                        posicion_47 = lista_limpia.index(elemento_a_buscar)
                        consumo_b = lista_limpia[posicion_47+1]
                        consumo = consumo_b.replace(' m3')
                    except:
                        print('elemento no se encuentra disponible')
                        consumo = 0
                    
                    #Posicion 48
                    texto_a_verificar = 'FACTOR DE COBRO'
                    posicion_48 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_48 = idx
                            break
                        
                    try:
                        factor_cobro = lista_limpia[posicion_48+1]
                    except:
                        print('elemento no se encuentra disponible')
                        factor_cobro = 0
                    
                    #Posicion 49
                    elemento_a_buscar = 'PERIODO'
                    try:
                        posicion_49 = lista_limpia.index(elemento_a_buscar)
                        periodo = lista_limpia[posicion_49+1]
                    except:
                        print('elemento no se encuentra disponible')
                        periodo = '-'
                    
                    #Posicion 50
                    elemento_a_buscar = 'LIMITE SOBRECONSUMO'
                    try:
                        posicion_50 = lista_limpia.index(elemento_a_buscar)
                        limite_sobreconsumo_b = lista_limpia[posicion_50+1]
                        limite_sobreconsumo = limite_sobreconsumo_b.replace(' m3')
                    except:
                        print('elemento no se encuentra disponible')
                        limite_sobreconsumo = 0

                #Cargamos libro excel donde volcaremos los datos
                libro = load_workbook(output_path+'/'+'Formato Planilla.xlsx')
                hoja_agua = libro['Agua']
                    
                ultima_fila = hoja_agua.max_row
                
                #Los datos mas importantes
                hoja_agua.cell(row=ultima_fila+1,column=1).value = 1
                hoja_agua.cell(row=ultima_fila+1,column=2).value = id_servicio
                
                #Primera tabla traspasada a excel
                #hoja_agua.cell(row=ultima_fila+1,column=1).value = factura_elec
                #hoja_agua.cell(row=ultima_fila+1,column=1).value = direccion_instalacion
                #hoja_agua.cell(row=ultima_fila+1,column=1).value = giro
                #hoja_agua.cell(row=ultima_fila+1,column=1).value = id_servicio
                hoja_agua.cell(row=ultima_fila+1,column=10).value = fecha_emision
                hoja_agua.cell(row=ultima_fila+1,column=11).value = fecha_vencimiento
                hoja_agua.cell(row=ultima_fila+1,column=12).value = direccion_central
                
                #Segunda tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=14).value = int(total_cargo_fijo)
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=18).value = int(cantidad_agua_potable)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=18).value = cantidad_agua_potable
                    
                hoja_agua.cell(row=ultima_fila+1,column=19).value = precio_agua_potable
                hoja_agua.cell(row=ultima_fila+1,column=20).value = int(total_agua_potable)
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=21).value = int(cantidad_sobreconsumo)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=21).value = cantidad_sobreconsumo
                    
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=22).value = int(precio_sobreconsumo)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=22).value = precio_sobreconsumo
                    
                try:
                    hoja_agua.cell(row=ultima_fila+1,column=23).value = int(total_sobreconsumo)
                except:
                    hoja_agua.cell(row=ultima_fila+1,column=23).value = total_sobreconsumo
                    
                hoja_agua.cell(row=ultima_fila+1,column=25).value =  int(cantidad_alcantarillado)
                hoja_agua.cell(row=ultima_fila+1,column=26).value = precio_alcantarillado
                #hoja_agua.cell(row=ultima_fila+1,column=23).value = total_alcantarillado   #Se repite
                hoja_agua.cell(row=ultima_fila+1,column=28).value =  int(cantidad_tratam_aguas_servidas)
                hoja_agua.cell(row=ultima_fila+1,column=29).value = precio_tratam_aguas_servidas
                hoja_agua.cell(row=ultima_fila+1,column=30).value = int(total_tratam_aguas_servidas)
                
                #Tercera tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=35).value = fecha_tarifa_publicada
                hoja_agua.cell(row=ultima_fila+1,column=36).value = cargo_variable_AP
                hoja_agua.cell(row=ultima_fila+1,column=37).value = cargo_variable_AP_punta
                hoja_agua.cell(row=ultima_fila+1,column=38).value = cargo_variable_AL
                hoja_agua.cell(row=ultima_fila+1,column=39).value = cargo_variable_sobreconsumo
                hoja_agua.cell(row=ultima_fila+1,column=40).value = int(corte_llave_paso)
                hoja_agua.cell(row=ultima_fila+1,column=41).value =  int(reposicion_llave)
                
                #Cuarta tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=42).value = int(monto_neto)
                hoja_agua.cell(row=ultima_fila+1,column=43).value = int(iva)
                hoja_agua.cell(row=ultima_fila+1,column=44).value =  int(total_mes)
                hoja_agua.cell(row=ultima_fila+1,column=45).value = int(saldo_anterior)
                hoja_agua.cell(row=ultima_fila+1,column=46).value = int(total_a_pagar)
                
                #Quinta tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=47).value = ruta
                hoja_agua.cell(row=ultima_fila+1,column=48).value = tipo_servicio
                hoja_agua.cell(row=ultima_fila+1,column=49).value = suministro
                hoja_agua.cell(row=ultima_fila+1,column=50).value =  grupo_tarifario
                hoja_agua.cell(row=ultima_fila+1,column=51).value =  medidor_general
                hoja_agua.cell(row=ultima_fila+1,column=52).value =  n_medidor
                hoja_agua.cell(row=ultima_fila+1,column=53).value =  int(diametro)
                hoja_agua.cell(row=ultima_fila+1,column=54).value = proxima_lectura
                hoja_agua.cell(row=ultima_fila+1,column=55).value = metodo_consumo
                hoja_agua.cell(row=ultima_fila+1,column=56).value = int(lector_actual_valor)
                hoja_agua.cell(row=ultima_fila+1,column=57).value = lector_actual_fecha
                hoja_agua.cell(row=ultima_fila+1,column=65).value = int(lectura_anterior_valor_real)
                hoja_agua.cell(row=ultima_fila+1,column=58).value = lector_anterior_fecha
                hoja_agua.cell(row=ultima_fila+1,column=59).value = clave_actual
                hoja_agua.cell(row=ultima_fila+1,column=60).value = int(consumo)
                hoja_agua.cell(row=ultima_fila+1,column=61).value = int(factor_cobro)
                hoja_agua.cell(row=ultima_fila+1,column=62).value = periodo
                hoja_agua.cell(row=ultima_fila+1,column=63).value = int(limite_sobreconsumo)

                libro.save(output_path+'/'+'Formato Planilla.xlsx')
                    
                #Copiamos el archivo a la carpeta outpu con el nombre que corresponde
                shutil.copy(archivo, output_path+nombre_oficial)
                print('-----')

        # #Obtenemos los archivos de la carpeta input
        # archivos_en_carpeta = os.listdir(folder_path)

        # # Iterar sobre los archivos y eliminarlos
        # for archivo in archivos_en_carpeta:
        #     ruta_archivo = os.path.join(folder_path, archivo)
        #     if os.path.isfile(ruta_archivo):
        #         os.remove(ruta_archivo)  
                    
                    