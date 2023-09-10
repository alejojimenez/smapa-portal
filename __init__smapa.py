from codigo.app_smapa import Scraper_Smapa
#import smtplib
from openpyxl import load_workbook

def send_notification():
    # Código para enviar correo electrónico de notificación
    print('')

if __name__ == '__main__':

    print('Obteniendo credenciales...')
    print('----------------------------------------------------------------------')
        
    credencials = '.\config\credenciales.xlsx'
    libro_accesos = load_workbook(credencials)
    hoja_credenciales = libro_accesos['Hoja1']
        
    for j in hoja_credenciales.iter_rows(2):
        try:
            rut = j[3].value
            passw = j[1].value
            web = j[2].value
            break
        except:
            ('no hay credenciales')
            
    email = rut
    password = passw
    url = web
    driver_path = 'chromedriver.exe'

    scraper = Scraper_Smapa(url, email, password, driver_path)
    print('ingresamos en la clase Scraper_Smapa...')
    print('----------------------------------------------------------------------')
    
    #Primera sociedad
    scraper.login()
    print('Hacemos login primera parte...')
    print('----------------------------------------------------------------------')
        
    scraper.scrapping_smapa(sociedad=1,limite=5)
    print('Hacemos scrapping...')
    print('----------------------------------------------------------------------')
    
    scraper.close()
    print('cerramos el bot 1era parte...')
    print('----------------------------------------------------------------------')
    
    #Segunda sociedad
    scraper.login()
    print('Hacemos login segunda parte...')
    print('----------------------------------------------------------------------')
        
    scraper.scrapping_smapa(sociedad=5,limite=8)
    print('Hacemos scrapping...')
    print('----------------------------------------------------------------------')

    scraper.close()
    print('cerramos el bot 2da parte...')
    print('----------------------------------------------------------------------')
    
    scraper.archivos()
    print('Extraemos datos...')
    print('----------------------------------------------------------------------')